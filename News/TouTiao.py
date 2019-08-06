import requests
import json
from openpyxl import Workbook
import time
import hashlib
import os
import datetime

# start_url获取方式：charles抓包获取
start_url = 'https://www.toutiao.com/api/pc/feed/?category=news_hot&utm_source=toutiao&widen=1&max_behot_time='
url = 'https://www.toutiao.com'

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'
}
cookies = {'tt_webid': '6720464891939669512'}

# 链接参数
max_behot_time = '0'
# 存储新闻标题
title = []
# 存储新闻的链接
source_url = []
# 存储新闻的完整链接
s_url = []
# 存储发布新闻的公众号
source = []
# 存储公众号的完整链接
media_url = {}


def get_as_cp():
    """
    该函数主要是为了获取as和cp参数，程序参考今日头条中的加密js文件：home_4abea46.js
    :return:
    """
    zz = {}
    # 获取当前计算机时间
    now = round(time.time())
    print(now)
    # hex()转换一个整数对象为16进制的字符串表示
    e = hex(int(now)).upper()[2:]
    print('e:', e)
    # hashlib.md5().hexdigest()创建hash对象并返回16进制结果
    a = hashlib.md5()
    print('a:', a)
    a.update(str(int(now)).encode('utf-8'))
    i = a.hexdigest().upper()
    print('i:', i)
    if len(e) != 8:
        # as,cp值抓包获取：Contents -> Query String
        zz = {'as': 'A1F5FD843951F95',
              'cp': '5D49C14FB9353E1'}
        return zz
    n = i[:5]
    a = i[-5:]
    r = ''
    s = ''
    for i in range(5):
        s = s + n[i] + e[i]
    for j in range(5):
        r = r + e[j + 3] + a[j]
    zz = {
        'as': 'A1' + s + e[-3:],
        'cp': e[0:3] + r + 'E1'
    }
    print('zz:', zz)
    return zz


def getdata(url, headers, cookies):
    """
    解析网页函数
    :param url:
    :param headers:
    :param cookies:
    :return:
    """
    r = requests.get(url, headers=headers, cookies=cookies)
    print(url)
    data = json.loads(r.text)
    return data


def savedata(title, s_url, source, media_url):
    """
    存储数据到文件
    :param title:
    :param s_url:
    :param source:
    :param media_url:
    :return:
    """
    # 存储数据到xlxs文件
    wb = Workbook()
    # 判断文件夹是否存在
    if not os.path.isdir(os.getcwd() + '/result'):
        # 新建存储文件夹
        os.makedirs(os.getcwd() + '/result')
    # 新建存储结果的excel文件
    filename = os.getcwd() + '/result/result-' + datetime.datetime.now().strftime(
        '%Y-%m-%d-%H-%m') + '.xlsx'
    # 新建的工作簿默认预先建好工作表，通过active属性获取
    ws = wb.active
    # 更改工作表的标题
    ws.title = 'data'
    # 对表格加入标题等
    ws['A1'] = '标题'
    ws['B1'] = '新闻链接'
    ws['C1'] = '头条号'
    ws['D1'] = '头条号链接'
    # 将数据写入表格
    for row in range(2, len(title) + 2):
        _ = ws.cell(column=1, row=row, value=title[row - 2])
        _ = ws.cell(column=2, row=row, value=s_url[row - 2])
        _ = ws.cell(column=3, row=row, value=source[row - 2])
        _ = ws.cell(column=4, row=row, value=media_url[source[row - 2]])
    # 保存文件
    wb.save(filename=filename)


def main(max_behot_time, title, source_url, s_url, source, media_url):
    """
    主函数
    :param max_behot_time:
    :param title:
    :param source_url:
    :param s_url:
    :param source:
    :param media_url:
    :return:
    """
    # 此处的数字类似于你刷新新闻的次数，正常情况下刷新一次会出现10条新闻，但也存在少于10条的情况；所以最后的结果并不一定是10的倍数
    for i in range(1):
        # 获取as和cp参数的函数
        ascp = get_as_cp()
        # https://www.toutiao.com/api/pc/feed/?min_behot_time=0&category=__all__&
        # utm_source=toutiao&widen=1&tadrequire=true&as=A1F5FD843951F95&cp=5D49C14FB9353E1&_signature=eC7RXxAfJWpxYqtKLpC5Pngu0U
        demo = getdata(
            start_url + max_behot_time + '&max_behot_time_tmp=' + max_behot_time + '&tadrequire=true&as=' + ascp[
                'as'] + '&cp=' + ascp['cp'], headers, cookies)

        print(demo)
        # time.sleep(1)
        for j in range(len(demo['data'])):
            # print(demo['data'][j]['title'])
            if demo['data'][j]['title'] not in title:
                # 获取新闻标题
                title.append(demo['data'][j]['title'])
                # 获取新闻链接
                source_url.append(demo['data'][j]['source_url'])
                # 获取发布新闻的公众号
                source.append(demo['data'][j]['source'])
            if demo['data'][j]['source'] not in media_url:
                # 获取公众号链接
                media_url[demo['data'][j]['source']] = url + demo['data'][j]['media_url']
        print(max_behot_time)
        # 获取下一个链接的max_behot_time参数的值
        max_behot_time = str(demo['next']['max_behot_time'])
        for index in range(len(title)):
            print('标题：', title[index])
            if 'https' not in source_url[index]:
                s_url.append(url + source_url[index])
                print('新闻链接：', url + source_url[index])
            else:
                print('新闻链接：', source_url[index])
                s_url.append(source_url[index])
                # print('源链接：', url+source_url[index])
            print('头条号：', source[index])
            print(len(title))  # 获取的新闻数量


if __name__ == '__main__':
    main(max_behot_time, title, source_url, s_url, source, media_url)
    savedata(title, s_url, source, media_url)